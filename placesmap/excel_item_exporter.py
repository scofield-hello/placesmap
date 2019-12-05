from logging import getLogger

from scrapy.exporters import BaseItemExporter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import BLACK
from collections import OrderedDict


class ExcelItemExporter(BaseItemExporter):
    """
    导出为Excel
    """
    _logger = getLogger('ExcelItemExporter')

    def __init__(self, filename: str, sheet_columns: dict, **kwargs):
        self._configure(kwargs, dont_fail=True)
        self._filename = filename
        self._workbook = Workbook()
        self._header_font = Font(name='宋体', bold=True, color=BLACK, size=11)
        self._header_alignment = Alignment(horizontal="center",
                                           vertical="center")
        self._normal_font = Font(name='宋体', bold=False, color=BLACK, size=11)
        self._normal_alignment = Alignment(horizontal="left",
                                           vertical="center")
        active_sheet = self._workbook.get_active_sheet()
        self._workbook.remove(active_sheet)
        self._workbook_sheets = []
        for (sheet_name, columns) in sheet_columns.items():
            worksheet = self._workbook.create_sheet(title=sheet_name)
            worksheet.sheet_properties.tabColor = '6FB7B7'
            worksheet.page_setup.fitToHeight = 1
            worksheet.page_setup.fitToWidth = 1
            worksheet.row_dimensions[1].height = 20
            worksheet.row_dimensions[2].height = 20
            worksheet.merge_cells(start_row=1,
                                  start_column=1,
                                  end_row=2,
                                  end_column=len(columns))
            worksheet_header_cell = worksheet['A1']
            worksheet_header_cell.value = sheet_name
            worksheet_header_cell.font = self._header_font
            worksheet_header_cell.alignment = self._header_alignment
            worksheet.row_dimensions[3].height = 25
            for row in worksheet.iter_rows(min_row=3,
                                           max_row=3,
                                           min_col=1,
                                           max_col=len(columns)):
                for cell in row:
                    cell.value = columns[cell.col_idx - 1]
                    cell.font = self._header_font
                    cell.alignment = self._header_alignment
                    worksheet.column_dimensions[cell.column].width = 30
            self._workbook_sheets.append(worksheet)
        self._logger.info('本次任务共创建%s个表格:[%s]' %
                          (len(self._workbook_sheets), self._workbook_sheets))

    def export_item(self, item):
        item_dict = dict(self._get_serialized_fields(item))
        serialized_item_dict = {}
        sheet_list = []
        for (key, value) in item_dict.items():
            if isinstance(value, str) or value is None:
                serialized_item_dict[key] = value
            elif isinstance(value, list):
                sheet_list.append(value)
            else:
                self._logger.warning("字段被忽略，字段: %s = %s, 类型: %s" %
                                     (key, value, type(value)))

        ordered_dict = OrderedDict(
            sorted(serialized_item_dict.items(), key=lambda t: int(t[0][1:])))
        sheet_list.insert(0, ordered_dict)
        for index, sheet_item in enumerate(sheet_list):
            if isinstance(sheet_item, dict):
                self._workbook_sheets[index].append(list(
                    ordered_dict.values()))
            elif isinstance(sheet_item, list):
                for data in sheet_item:
                    ordered_dict = OrderedDict(
                        sorted(data.items(), key=lambda t: int(t[0][1:])))
                    self._workbook_sheets[index].append(
                        list(ordered_dict.values()))

    def finish_exporting(self):
        self._logger.info('当前任务结束, 文档保存地址为:%s' % self._filename)
        self._workbook.save(self._filename)
